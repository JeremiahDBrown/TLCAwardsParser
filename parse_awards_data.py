# -*- coding: utf-8 -*-
"""
Created on Sun Dec 15 22:01:45 2024

@author: JBrown
"""
import os
from bs4 import BeautifulSoup
import pickle
from datetime import datetime
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter.simpledialog import askstring
from pathlib import Path

class COH_Report:
    def __init__(self, infile=None):
        self.infile = infile
        self.previous_awards = []
        self.new_awards = []
        self.date = None
        self.reset_award_count()
        self.new_html_names = []
        
        if infile is not None:
            if 'pkl' in os.path.splitext(infile)[-1].lower():
                self.load_data(infile)            
            else:
                self.parse_court_of_honor_html(infile)
        
    @staticmethod
    def merge_awards(list1, list2):
        def dict_key(d):
            return tuple(d.get(k) for k in ['award_name', 'additional_data', 'completion_date'])
        seen = {dict_key(d) for d in list1}
        merged = list1[:]
        for d in list2:
            if dict_key(d) not in seen:
                merged.append(d)
                # seen.add(dict_key(d))
        return merged

    @staticmethod
    def combine_lists(list1, list2):
        combined_recs = []
        if len(list1) == 0:
            combined_recs = list2.copy()
        elif len(list2) == 0:
            combined_recs = list1.copy()
        else:
            for list1_val in list1:
                record = list1_val.copy()
                rec2 = next((item for item in list2 if item.get('name') == record['name']), None)
                if rec2 is not None:
                    record['awards'] = COH_Report.merge_awards(record['awards'], rec2['awards'])
                combined_recs.append(record)
            for record in list2:
                rec1 = next((item for item in list1 if item.get('name') == record['name']), None)
                if rec1 is None:
                    combined_recs.append(record)
        return combined_recs
        
        
    def reset_award_count(self):
        keys = ['Worthy Life Award','Worthy Life Cross','Ceremonial Standard','Standard Medallion',
                'Fox Branch Patch','Fox Forest Award','Bronze Branch Pins','Bronze Sylvan Stars',
                'Hawk Branch Patch','Hawk Forest Award','Silver Branch Pins','Silver Sylvan Stars',
                'Mountain Lion Branch Patch','Mountain Lion Forest Award','Gold Branch Pins','Gold Sylvan Stars',
                'Fireguard','Woodsman','Timberline Award','Recruit Trailman Rank','Able Trailman Rank','Ready Trailman Rank',
                'Navigator Service Star','Ridgeline Award','Journey Rank','Ascent Rank','Horizon Rank','Adventurer Service Star',
                'Aquatics','Camping','Fire Ranger','First Aid','Our Flag','Outdoor Cooking','Ropework','Trail Skills','Woods Tools',
                'Citizenship','Cycling','Emergency Preparedness','Family Man','Fitness','Hiking','Outdoor Life','Personal Resources','Swimming',
                'Heritage Elective','Hobbies Elective','Life Skills Elective','Outdoor Skills Elective','Science and Technology Elective',
                'Sports and Fitness Elective','Values Elective']
        self.award_count = dict.fromkeys(keys,0)

    def get_matching_records(self, keyname, value, search_new=True, search_previous=False):
        # gets all awards records with keyname==value. keyname is typically name or program_level
        # If search_new is True, look for matches in new_awards list. If search_previous is True
        # look for matches in previous_awards list.
        # If both are false, get matches from new_awards list only if the name is contained in
        # new_html_names.
        matches = []
        if search_new:
            matches = [entry for entry in self.new_awards if entry.get(keyname) == value]
        if search_previous:
            matches_old = [entry for entry in self.previous_awards if entry.get(keyname) == value]
            if search_new:
                matches = COH_Report.combine_lists(matches,matches_old)
            else:
                matches = matches_old
        if not (search_new or search_previous):
            matches = [entry for entry in self.new_awards if 
                       (entry.get(keyname) == value) and (entry.get('name') in self.new_html_names)]
            
        return matches
        
    def save_data(self, datafilename=None):
        if datafilename is None:
            datafilename = 'COH-data-' + datetime.now().strftime("%Y%m%d%H%M") + '.pkl'
        with open(datafilename, 'wb') as outf:
            pickle.dump(self, outf)
        
    def load_data(self, datafile):
        with open(datafile, 'rb') as infile:
            old_data = pickle.load(infile)
            self.new_awards = COH_Report.combine_lists(self.new_awards, old_data.new_awards)
            if self.date is None or old_data.date < self.date:
                self.previous_awards = old_data.previous_awards
            if self.date is None:
                self.date = old_data.date
            else:
                self.date = max(self.date,old_data.date)
            self.generate_awards_program('')    # update awards counts
            if self.infile is None:
                self.infile = old_data.infile
    
    def parse_court_of_honor_html(self, infile, read_purchased_as_awarded=True):
        self.infile = infile
        with open(infile, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')
    
        previous_awards_list = []
        new_awards_list = []
        self.new_html_names = []
    
        # Find all panels corresponding to individual records
        panels = soup.find_all('div', class_='panel rounded shadow no-overflow')
    
        for panel in panels:
            record = {}
    
            # Extract name and program level
            header = panel.find('div', class_='profile_header')
            if header:
                name, program_level = None, None
                header_contents = header.find('div').contents  # Get all child nodes of the header
                for content in header_contents:
                    if isinstance(content, str):
                        # Extract plain text (name)
                        if content.strip():  # Ignore empty strings
                            name = " ".join([subname.strip() for subname in content.split(',',maxsplit=1)][::-1])

                    elif content.name == 'span':
                        # Extract the program level from <span> tag
                        program_level = content.text.strip()
    
                record['name'] = name
                record['program_level'] = program_level
    
            # Extract awards
            new_awards = []
            purchased_awards = []
            table = panel.find('table', class_='table-basic')
            if table:
                rows = table.find('tbody').find_all('tr')
                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) >= 6:
                        award_name = cols[1].find('strong').text.strip()
                        additional_data = cols[1].find('i', class_='faded-style')
                        additional_data = additional_data.text.strip() if additional_data else None
                        completion_date = cols[2].text.strip()
                        purchased = bool(cols[3].find('i', class_='fas fa-lg fa-check')) and read_purchased_as_awarded
                        awarded = bool(cols[4].find('i', class_='fas fa-lg fa-check'))
                        awarded_date = cols[5].find('input', class_='krajee-datepicker')
                        awarded_date = awarded_date['value'].strip() if awarded_date else None
                        
                        if 'Electives (' in additional_data:
                            frontier = re.split(r'[()]', additional_data)[1].strip()
                            additional_data = f'{frontier} Elective'
                        if 'Design Your Own Badge' in award_name:
                            # askstring("DYOB Badge", "Enter name of badge earned by {ABC} on {XYZ}:")
                            award_name = input(
                                f'Please enter the correct name for the Design Your Own/TEAMS {additional_data} Badge earned by {name} on {completion_date}:'
                                )
                        if program_level in ['Navigator', 'Adventurer'] and 'Rank' in award_name and purchased:
                            purchased = False
                            additional_data = "Previously Announced"
                            print(f'{award_name} earned by {name} on {completion_date} was marked as purchased. Including in to-award list.')
                        
                        branch_color = None
                        if program_level in ['Fox', 'Hawk', 'Mountain Lion'] and (('Branch Pin' in award_name) or ('Sylvan Star' in award_name)):
                            match(award_name.split(maxsplit=1)[0].lower()):
                                case 'heritage':
                                    branch_color = 'Brown'
                                case 'life':
                                    branch_color = 'Burgundy'
                                case 'science':
                                    branch_color = 'Yellow'
                                case 'hobbies':
                                    branch_color = 'Black'
                                case 'values':
                                    branch_color = 'Red'
                                case 'sports':
                                    branch_color = 'Green'
                                case 'outdoor':
                                    branch_color = 'Blue'
                            if 'Branch Pin' in award_name:
                                additional_data = program_level + ' ' + branch_color + ' Branch Pin'
                            if 'Sylvan Star' in award_name:
                                additional_data = program_level + ' ' + branch_color + ' Sylvan Star'
                            # award_name = award_name.split('(')[0].strip()
                        
                        if purchased:
                            purchased_awards.append({
                                'award_name': award_name,
                                'additional_data': additional_data,
                                'completion_date': completion_date,
                                'purchased': purchased,
                                #'awarded': awarded,
                                #'awarded_date': awarded_date
                            })
                        else:
                            new_awards.append({
                                'award_name': award_name,
                                'additional_data': additional_data,
                                'completion_date': completion_date,
                                'purchased': purchased,
                                #'awarded': awarded,
                                #'awarded_date': awarded_date
                            })

            # Append the record to the list
            if len(new_awards) > 0:
                record['awards'] = new_awards
                new_awards_list.append(record)
                self.new_html_names.append(name)
            if len(purchased_awards) > 0:
                record_old = record.copy()
                record_old['awards'] = purchased_awards
                previous_awards_list.append(record_old)
        
        # integrate new data
        self.new_awards = COH_Report.combine_lists(self.new_awards, new_awards_list)
        filedate = datetime.strptime(awarded_date, '%m/%d/%Y').date()
        if self.date is None or filedate < self.date:
            self.previous_awards = previous_awards_list
        if self.date is None:
            self.date = filedate
        else:
            self.date = max(self.date,filedate)
        self.generate_awards_program('')    # update awards counts
        
    def generate_detailed_output(self, output_file=None, print_new=True, print_previous=False):
        # Generate html for award cards for each Trailman.
        # output_file: name of file to write. If None, generate filename including current timestamp.
        # print_new: Include all new_awards
        # print_previous: Include all previous_awards and mark them as purchased
        # If both print_new and print_previous are false, write awards from new_awards list 
        #   only if the Trailman name was just read in the html file.
        def person_block(person, mark_purchased):
            name = person["name"]
            level = person["program_level"]
            if mark_purchased:
                awards_rows = "".join(
                    f"<tr><td>{a['award_name']}</td><td>{a['additional_data']}</td><td>{a['completion_date']}</td><td>{'X' if a['purchased'] else ''}</td></tr>"
                    for a in person["awards"]
                )
                return f"""
                <div class="person-block">
                    <h3>{name} <span class="level">({level})</span></h3>
                    <table>
                        <thead>
                            <tr><th>Award</th><th>Details</th><th>Date</th><th>Purchased</th></tr>
                        </thead>
                        <tbody>
                            {awards_rows}
                        </tbody>
                    </table>
                </div>
                """
            else:
                awards_rows = "".join(
                    f"<tr><td>{a['award_name']}</td><td>{a['additional_data']}</td><td>{a['completion_date']}</td></tr>"
                    for a in person["awards"]
                )
                return f"""
                <div class="person-block">
                    <h3>{name} <span class="level">({level})</span></h3>
                    <table>
                        <thead>
                            <tr><th>Award</th><th>Details</th><th>Date</th></tr>
                        </thead>
                        <tbody>
                            {awards_rows}
                        </tbody>
                    </table>
                </div>
                """
    
        if output_file is None:
            output_file = 'COH-awards-report-' + datetime.now().strftime("%Y%m%d%H%M") + '.html'
        
        fox_blocks = "\n".join(person_block(p, print_previous) for p in self.get_matching_records('program_level', 'Fox', print_new, print_previous))
        hawk_blocks = "\n".join(person_block(p, print_previous) for p in self.get_matching_records('program_level', 'Hawk', print_new, print_previous))
        ml_blocks = "\n".join(person_block(p, print_previous) for p in self.get_matching_records('program_level', 'Mountain Lion', print_new, print_previous))
        nav_blocks = "\n".join(person_block(p, print_previous) for p in self.get_matching_records('program_level', 'Navigator', print_new, print_previous))
        adv_blocks = "\n".join(person_block(p, print_previous) for p in self.get_matching_records('program_level', 'Adventurer', print_new, print_previous))
    
        html_template = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: sans-serif;
                    margin: 0.5in;
                }}
                .container {{
                    column-count: 2;
                    column-gap: 0.5in;
                }}
                .person-block {{
                    break-inside: avoid;
                    margin-bottom: 1em;
                    border: 1px solid #000;
                    padding: 8px;
                    border-radius: 6px;
                    background: none;
                    font-size: 9pt;
                    width: 100%;
                }}
                table {{
                    width: 100%;
                    table-layout: fixed;
                    font-size: 8pt;
                    border-collapse: collapse;
                }}
                th, td {{
                    border: 1px solid #aaa;
                    padding: 2px 4px;
                    text-align: left;
                }}
                th {{
                    background-color: #eee;
                }}
                h3 {{
                    margin: 0 0 0.5em 0;
                }}
                .level {{
                    font-weight: normal;
                    font-size: 0.9em;
                    color: #666;
                }}
                @media print {{
                    .page-break {{
                        page-break-before: always;
                    }}
                }}
                @page {{
                    margin: 0.5in;
                }}
            </style>
        </head>
        <body>
            <div class="section">
                <h2>Fox Awards</h2>
                <div class="container">
                    {fox_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h2>Hawk Awards</h2>
                <div class="container">
                    {hawk_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h2>Mountain Lion Awards</h2>
                <div class="container">
                    {ml_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h2>Navigator Awards</h2>
                <div class="container">
                    {nav_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h2>Adventurer Awards</h2>
                <div class="container">
                    {adv_blocks}
                </div>
            </div>
        </body>
        </html>
        """
    
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(html_template)

    def generate_awards_program(self, output_file=None):
        # pass in '' for output_file to recalculate awards count without generating output file
        def wt_summary_block(person):
            name = person["name"]
            level = person["program_level"]
            match level:
                case 'Fox':
                    color = 'Bronze'
                case 'Hawk':
                    color = 'Silver'
                case 'Mountain Lion':
                    color = 'Gold'
            
            award_str = ''
            matches = [entry for entry in person["awards"] if 'Joining Award' in entry.get('award_name')]
            if len(matches) > 0:
                award_name = f'{level} Branch Patch'
                award_str = award_str + f'<tr><td>{award_name}</td></tr>'
                self.award_count[award_name] += 1
            matches = [entry for entry in person["awards"] if 'Forest Award' in entry.get('award_name')]
            if len(matches) > 0:
                award_name = f'{level} Forest Award'
                award_str = award_str + f'<tr><td>{award_name}</td></tr>'
                self.award_count[award_name] += 1
            matches = [entry for entry in person["awards"] if 'Branch Pin' in entry.get('award_name')]
            if len(matches) > 0:
                award_str = award_str + f'<tr><td>{len(matches)} {level} Branch Pins</td>'
                self.award_count[f'{color} Branch Pins'] += len(matches)
            matches = [entry for entry in person["awards"] if 'Sylvan Star' in entry.get('award_name')]
            if len(matches) > 0:
                award_str = award_str + f'<tr><td>{len(matches)} {level} Sylvan Star</td></tr>'
                self.award_count[f'{color} Sylvan Stars'] += len(matches)
            matches = [entry for entry in person["awards"] if 'Fireguard' in entry.get('award_name')]
            if len(matches) > 0:
                award_name = 'Fireguard'
                award_str = award_str + f'<tr><td>{award_name} Award</td></tr>'
                self.award_count[award_name] += 1
            matches = [entry for entry in person["awards"] if 'Woodsman' in entry.get('award_name')]
            if len(matches) > 0:
                award_name = 'Woodsman'
                award_str = award_str + f'<tr><td>{award_name} Award</td></tr>'
                self.award_count[award_name] += 1

            return f"""
            <div class="person-block">
                <h3>{name} <span class="level">({level})</span></h3>
                <table>
                    <thead>
                        <tr><th>Awards</th></tr>
                    </thead>
                    <tbody>
                        {award_str}
                    </tbody>
                </table>
            </div>
            """

        def navadv_summary_block(person):
            name = person["name"]
            level = person["program_level"]
            
            award_str = ''
            matches = [entry for entry in person["awards"] if ('Ready (' in entry.get('additional_data')) 
                       or ('Horizon (' in entry.get('additional_data'))]
            for badge in matches:
                award_name = badge['award_name'].split('(',maxsplit=1)[0].strip()
                award_str = award_str + f'<tr><td>{award_name}</td></tr>'
                self.award_count[award_name] += 1
            matches = [entry for entry in person["awards"] if ' Elective' in entry.get('additional_data')]
            for badge in matches:
                award_name = badge['award_name']
                award_str = award_str + f'<tr><td>{award_name}</td></tr>'
                self.award_count[badge['additional_data']] += 1
            matches = [entry for entry in person["awards"] if 'Navigator Service Star' in entry.get('award_name')]
            if len(matches)>0:
                award_str = award_str + f'<tr><td>{len(matches)} Navigator Service Stars</td></tr>'
                self.award_count['Navigator Service Star'] += len(matches)
            matches = [entry for entry in person["awards"] if 'Adventurer Service Star' in entry.get('award_name')]
            if len(matches)>0:
                award_str = award_str + f'<tr><td>{len(matches)} Adventurer Service Stars</td></tr>'
                self.award_count['Adventurer Service Star'] += len(matches)
            return f"""
            <div class="person-block">
                <h3>{name} <span class="level">({level})</span></h3>
                <table>
                    <thead>
                        <tr><th>Awards</th></tr>
                    </thead>
                    <tbody>
                        {award_str}
                    </tbody>
                </table>
            </div>
            """

        def rank_block(rank):
            award_str = ''
            for p in self.new_awards:
                matches = [entry for entry in p["awards"] if rank in entry.get('award_name')]
                if len(matches)>0:
                    award_str = award_str + f'<tr><td>{p["name"]}</td></tr>'
                    self.award_count[rank] += 1
                    if 'Able' in rank:
                        self.award_count['Ceremonial Standard'] += 1
                        self.award_count['Standard Medallion'] += 1
            if len(award_str) == 0:
                return ""
            else:
                return f"""
                <div class="person-block">
                    <h3>{rank}</h3>
                    <table>
                        <thead>
                            <tr><th>Name</th></tr>
                        </thead>
                        <tbody>
                            {award_str}
                        </tbody>
                    </table>
                </div>
                """
            
        def worthylife_block(levels):
            award_str = ''
            for p in self.new_awards:
                matches = [entry for entry in p["awards"] if 'Worthy Life Award' in entry.get('award_name')]
                if len(matches)>0 and p["program_level"] in levels:
                    award_str = award_str + f'<tr><td>{p["name"]}</td><td>{p["program_level"]}</td></tr>'
                    self.award_count['Worthy Life Award'] += 1
                    self.award_count['Worthy Life Cross'] += 1
            return f"""
            <div class="person-block">
                <h3>Worthy Life Awards</h3>
                <table>
                    <thead>
                        <tr><th>Name</th><th>Level</th></tr>
                    </thead>
                    <tbody>
                        {award_str}
                    </tbody>
                </table>
            </div>
            """
            
        def capstone_block(award_name):
            award_str = ''
            for p in self.new_awards:
                matches = [entry for entry in p["awards"] if award_name in entry.get('award_name')]
                if len(matches)>0:
                    award_str = award_str + f'<tr><td>{p["name"]}</td></tr>'
                    self.award_count[award_name] += 1
            return f"""
            <div class="person-block">
                <h3>{award_name}s</h3>
                <table>
                    <thead>
                        <tr><th>Name</th></tr>
                    </thead>
                    <tbody>
                        {award_str}
                    </tbody>
                </table>
            </div>
            """
            
        if output_file is None:
            output_file = 'COH-awards-program-' + datetime.now().strftime("%Y%m%d%H%M") + '.html'
        
        self.reset_award_count()
        fox_blocks = "\n".join(wt_summary_block(p) for p in self.get_matching_records('program_level', 'Fox'))
        hawk_blocks = "\n".join(wt_summary_block(p) for p in self.get_matching_records('program_level', 'Hawk'))
        ml_blocks = "\n".join(wt_summary_block(p) for p in self.get_matching_records('program_level', 'Mountain Lion'))

        wt_special_blocks = "\n".join([worthylife_block(['Fox','Hawk','Mountain Lion']), capstone_block('Timberline Award')])
        
        nav_blocks = "\n".join(navadv_summary_block(p) for p in self.get_matching_records('program_level', 'Navigator'))
        adv_blocks = "\n".join(navadv_summary_block(p) for p in self.get_matching_records('program_level', 'Adventurer'))
        rank_adv_blocks = "\n".join(rank_block(p) for p in self.award_count.keys() if 'Rank' in p)
        navadv_special_blocks = "\n".join([worthylife_block(['Navigator','Adventurer']), capstone_block('Ridgeline Award')])
        
        html_template = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: sans-serif;
                    margin: 0.5in;
                }}
                .container {{
                    column-count: 3;
                    column-gap: 0.5in;
                }}
                .person-block {{
                    break-inside: avoid;
                    margin-bottom: 1em;
                    border: 1px solid #000;
                    padding: 8px;
                    background: none;
                    font-size: 10pt;
                    width: 100%;
                }}
                table {{
                    width: 100%;
                    table-layout: fixed;
                    font-size: 9pt;
                    border-collapse: collapse;
                }}
                th, td {{
                    border: 1px solid #aaa;
                    padding: 2px 4px;
                    text-align: left;
                }}
                th {{
                    background-color: #eee;
                }}
                h3 {{
                    margin: 0 0 0.5em 0;
                }}
                .level {{
                    font-weight: normal;
                    font-size: 0.9em;
                    color: #666;
                }}
                @media print {{
                    .page-break {{
                        page-break-before: always;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="section">
                <h1>Fox Awards</h2>
                <div class="container">
                    {fox_blocks}
                </div>
            </div>

            <div class="section">
                <h1>Hawk Awards</h2>
                <div class="container">
                    {hawk_blocks}
                </div>
            </div>

            <div class="section">
                <h1>Mountain Lion Awards</h2>
                <div class="container">
                    {ml_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h1>Woodland Trails Special Awards</h2>
                <div class="container">
                    {wt_special_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h1>Navigator Awards</h2>
                <div class="container">
                    {nav_blocks}
                </div>
            </div>

            <div class="section">
                <h1>Adventurer Awards</h2>
                <div class="container">
                    {adv_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h1>Rank Advancement Awards</h2>
                <div class="container">
                    {rank_adv_blocks}
                </div>
            </div>

            <div class="section page-break">
                <h1>Navigators/Adventurers Special Awards</h2>
                <div class="container">
                    {navadv_special_blocks}
                </div>
            </div>
        </body>
        </html>
        """
    
        if len(output_file) > 0:
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(html_template)

    def generate_shopping_list(self, output_file=None):
        if output_file is None:
            output_file = 'COH-shopping-list-' + datetime.now().strftime("%Y%m%d%H%M") + '.xlsx'
        skip_rows = [5,6,22,23,29,30,35,36,55,56]
        extra_comment_rows = [1,24,25,26,31,32,33]
        extra_comments = ['* Some Worthy Life Awards may only requires Crosses',
                          '* May have already been awarded at a meeting.  Announce at COH.',
                          '* Patch may have already been awarded at a meeting.  Award standard, standard medallion, and Able medallion.',
                          '* Patch may have already been awarded at a meeting.  Award Ready medallion.',
                          '* Patch may have already been awarded at a meeting.  Award Journey medallion.',
                          '* Patch may have already been awarded at a meeting.  Award Ascent medallion.',
                          '* Patch may have already been awarded at a meeting.  Award Horizon medallion.',
                          ]
        row_offset = 2
        name_col = 2
        count_col = 4
        comment_col = 6
        wb = Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%b%Y") + " COH Shopping List"
        ws.column_dimensions[get_column_letter(name_col)].width = 31.03
        ws.column_dimensions[get_column_letter(count_col)].width = 12.9
        ws.column_dimensions[get_column_letter(comment_col)].width = 12.9
        ws.row_dimensions[1].height = 30
        ws.cell(1,count_col).alignment = Alignment(horizontal='center',vertical='center')
        ws.cell(1,count_col,'To Award')
        
        row_colors = {}
        for r in range(1,3):
            row_colors[r] = "D9D9D9"
        for r in list(range(3, 5)) + list(range(57, 64)):
            row_colors[r] = "70AD47"
        for r in range(7,11):
            row_colors[r] = "A9D08E"
        for r in range(11,15):
            row_colors[r] = "FFD966"
        for r in range(15,22):
            row_colors[r] = "9BC2E6"
        for r in list(range(24, 29)) + list(range(37, 46)):
            row_colors[r] = "F4B084"
        for r in list(range(31, 35)) + list(range(46, 55)):
            row_colors[r] = "8EA9DB"
        
        thinbord = Side(border_style="thin", color="000000")
        medbord = Side(border_style="medium", color="000000")
        center_align = Alignment(horizontal='center')
        
        cur_row = 1
        for award_name, count in self.award_count.items():
            patfill = PatternFill(start_color=row_colors[cur_row], end_color=row_colors[cur_row], fill_type="solid")
            
            if cur_row in [1,7,24,31,37,57]:
                border = Border(top=medbord,bottom=thinbord,left=medbord,right=thinbord)
            elif cur_row in [4,10,14,21,28,34,45,54,63]:
                border = Border(top=thinbord,bottom=medbord,left=medbord,right=thinbord)
            else:
                border = Border(top=thinbord,bottom=thinbord,left=medbord,right=thinbord)
            ws.cell(cur_row+row_offset, name_col, award_name)
            ws.cell(cur_row+row_offset, name_col).fill = patfill
            ws.cell(cur_row+row_offset, name_col).border = border
            
            ws.cell(cur_row+row_offset, count_col, count)
            ws.cell(cur_row+row_offset, count_col).fill = patfill
            ws.cell(cur_row+row_offset, count_col).border = border
            
            if cur_row in extra_comment_rows:
                ws.cell(cur_row+row_offset, comment_col, extra_comments[extra_comment_rows.index(cur_row)])
            cur_row += 1
            while cur_row in skip_rows:
                cur_row += 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=count_col).alignment = center_align
        wb.save(output_file)

def latest_data_file():
    pattern = re.compile(r'COH-data-(\d{12})\.pkl')
    files = glob.glob('COH-data-*.pkl')
    most_recent_datafile = []
    most_recent_date = datetime(2014,1,1)
    for file in files:
        dmatch = pattern.search(file)
        if dmatch:
            dt = datetime.strptime(dmatch.group(1),'%Y%m%d%H%M')
            if (most_recent_date is None) or (dt > most_recent_date):
                most_recent_date = dt
                most_recent_datafile = file
    return most_recent_datafile
    
        
# Example usage - Set initial_dump to true normally. If you have updates to awards already processed,
# save the new CoH Report to 'AdditionalAwards.htm' and set initial_dump to False 
initial_dump = True

Tk().withdraw()
input_file_path = askopenfilename(title="Court of Honor Report File", filetypes=[("HTML Files",("*.htm","*.html"))])
file_path = Path(input_file_path).resolve().parent
os.chdir(file_path)

awards_program = 'awards_program.html'
award_cards = 'awards_detail_cards.html'
if initial_dump:
    full_dump_file = 'awards_full_dump.html'
else:
    full_dump_file = 'updated_awards_full_dump.html'
    previous_data_file = latest_data_file()
    new_awards_shopping_list = 'temp_new_awards_shopping_list.xlsx'
    updated_award_cards = 'updated_awards_detail_cards.html'
       
cohdata = COH_Report(input_file_path)
if not initial_dump:
    cohdata.generate_shopping_list(new_awards_shopping_list)
    cohdata.load_data(previous_data_file)
    cohdata.generate_detailed_output(updated_award_cards, False, False)

cohdata.generate_detailed_output(full_dump_file,print_previous=True)
cohdata.generate_detailed_output(award_cards)
cohdata.generate_awards_program(awards_program)
cohdata.generate_shopping_list()
cohdata.save_data()
